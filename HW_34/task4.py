import openpyxl
from datetime import datetime

class Test1Results:
    # Class attribute for the deadline
    deadline = "2024-10-28 23:59:59"

    def __init__(self, student_id, file_path="test_1.xlsx"):
        """
        Initialize with the Excel file and student ID (ИСУ).

        :param file_path: path to the .xlsx file
        :param student_id: student ID (номер ИСУ)
        """
        self.file_path = file_path
        self.student_id = student_id
        self.data = self._load_data()  # метод для загрузки данных из файла
        # your code here
        #хранит конкретный словарь по заданному ИСУ
        found_dict = next((d for d in self.data if d["ИСУ"] == self.student_id), None)

        if found_dict is not None:
            self.grade = found_dict.get('grade')  # оценка студента с номером ИСУ student_id
            self.timestamp = found_dict.get('timestamp')  # время сдачи теста студента с номером ИСУ student_id
        else:
            self.grade = None
            self.timestamp = None
            print(f"Student with ID {self.student_id} not found.")

    def _load_data(self):
        data = []
        wb_obj = openpyxl.load_workbook(self.file_path)
        sheet_obj = wb_obj.active

        def datt(row_num, col_num):
            return sheet_obj.cell(row=row_num, column=col_num).value

        def str_to_timestamp(row_num):
            cell_obj = sheet_obj.cell(row=row_num, column=3).value
            try:
                return datetime.strptime(str(cell_obj), "%Y-%m-%d %H:%M:%S")
            except (ValueError, TypeError):
                print('СОС')

        def grade_counter(row_num):
            total = 0
            for i in range(4, 23):
                cell_obj = sheet_obj.cell(row=row_num, column=i).value
                if cell_obj:
                    total += int(cell_obj)
            return total

        def constructor():
            for i in range(2, 117):
                my_dict = {
                    "ФИО": datt(row_num=i , col_num=1),
                    "ИСУ": datt(row_num= i, col_num= 2),
                    "timestamp": str_to_timestamp(i),
                    "grade": grade_counter(i),
                }
                data.append(my_dict)

        constructor()
        # print(data[0])
        return data

    def is_late(self):
            """
            Check if the student's submission was late.

            :returns: True if self.timestamp > is past the Test1Results.deadline or False otherwise
            """

            # your code here

            deadline_datetime =  datetime.strptime(Test1Results.deadline,"%Y-%m-%d %H:%M:%S")
            submission_datetime = self.timestamp

            result = submission_datetime > deadline_datetime
            print(result)


# Пример использования
if __name__ == "__main__":
    test_results = Test1Results(468130,"test_1.xlsx")  # попробуйте любой номер ИСУ из доступных в таблице

    print(f"Grade for student {test_results.student_id}: {test_results.grade}")

    is_late = test_results.is_late()
    print(f"Was the submission late? {'Yes' if is_late else 'No'}")